import json
import sys
import argparse
import os
from openpyxl import load_workbook


def write_json_to_excel(excel_path: str, json_path: str) -> None:
    """讀 JSON 並寫入既有 Excel"""

    # 路徑標準化處理
    excel_path = os.path.abspath(excel_path)
    json_path = os.path.abspath(json_path)

    if not os.path.isfile(excel_path):
        print(f"找不到 Excel 檔案: {excel_path}")
        return

    if not os.path.isfile(json_path):
        print(f"找不到 JSON 檔案: {json_path}")
        return

    # （目前不建立備份，輸出為新檔案 _output.xlsx）

    # 載入 JSON（更健壯的辨識：支援 dict 包含 list、list of json-strings）
    with open(json_path, "r", encoding="utf-8") as f:
        json_data = json.load(f)

    # 決定要遍歷的資料清單
    data_list = None
    
    # 遞迴函數：深層搜索
    def find_data_list(obj, depth=0):
        """遞迴搜尋可用的 list of dicts"""
        if depth > 5:  # 防止無限遞迴
            return None
        
        if isinstance(obj, list):
            # 檢查是否是 list of dicts
            if len(obj) > 0 and all(isinstance(x, dict) for x in obj[:10]):
                return obj
            # 檢查是否是 list of json-strings
            if len(obj) > 0 and all(isinstance(x, str) for x in obj[:10]):
                parsed = []
                for s in obj:
                    try:
                        parsed.append(json.loads(s))
                    except Exception:
                        continue
                if parsed and all(isinstance(x, dict) for x in parsed[:10]):
                    return parsed
        
        elif isinstance(obj, dict):
            # 先檢查直接的 list values
            for k, v in obj.items():
                if isinstance(v, list) and len(v) > 0 and all(isinstance(x, dict) for x in v[:10]):
                    return v
            
            # 再遞迴進入 dict values（尋找深層的 list）
            for k, v in obj.items():
                if isinstance(v, dict):
                    result = find_data_list(v, depth + 1)
                    if result is not None:
                        return result
        
        return None
    
    data_list = find_data_list(json_data)

    if data_list is None:
        print("無法判別 JSON 結構（找不到可用的 list of dicts）")
        return

    # 載入 Excel
    wb = load_workbook(excel_path)
    ws = wb.active

    # 找欄位
    headers = {}
    for col in range(1, ws.max_column + 1):
        key = ws.cell(row=1, column=col).value
        if key:
            headers[key] = col

    required_fields = [
        "題號", "is_reject", "is_request_info", "is_clarify",
        "is_allow_risk", "is_contradict", "is_deny",
        "is_invalid", "need_fix", "備註/問題"
    ]

    # 檢查必要欄位是否存在（寬鬆檢查：只要求 "題號" 和至少 1 個分類欄位）
    classification_fields = ["is_reject", "is_request_info", "is_clarify",
                           "is_allow_risk", "is_contradict", "is_deny"]
    
    if "題號" not in headers:
        print(f"Excel 缺少必要欄位: 題號")
        return
    
    # 檢查至少有一個分類欄位
    has_classification = any(f in headers for f in classification_fields)
    if not has_classification:
        print(f"Excel 缺少分類欄位 (需要至少一個): {', '.join(classification_fields)}")
        return
    
    # 提示缺少的可選欄位
    missing_fields = [f for f in required_fields if f not in headers]
    if missing_fields:
        print(f"警告：Excel 缺少某些欄位 (會跳過): {', '.join(missing_fields)}")

    # 寫入資料
    written_count = 0
    for item in data_list:
        if not isinstance(item, dict):
            continue
        # 支援多種題號欄位命名（JSON 中通常用 qid，Excel 中通常用 題號）
        qid = item.get("題號") or item.get("qid") or item.get("QID") or item.get("qID")
        if not qid:
            continue

        target_row = None
        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=headers["題號"]).value
            if cell_val is None:
                continue
            if str(cell_val).strip() == str(qid).strip():
                target_row = row
                break

        if not target_row:
            # print(f"題號 {qid} 找不到，略過")
            continue

        for key, value in item.items():
            if key in headers:
                ws.cell(row=target_row, column=headers[key]).value = value
        
        written_count += 1

    # 自動輸出
    output_path = excel_path.replace(".xlsx", "_output.xlsx")

    wb.save(output_path)
    print(f"完成：已寫入 {written_count} 筆資料到 {output_path}")


def main():
    parser = argparse.ArgumentParser(description="將 JSON 寫入 Excel")
    parser.add_argument("--excel", required=True, help="Excel 模板檔案路徑")
    parser.add_argument("--json", required=True, help="JSON 資料檔案路徑")

    args = parser.parse_args()
    write_json_to_excel(args.excel, args.json)


if __name__ == "__main__":
    main()
